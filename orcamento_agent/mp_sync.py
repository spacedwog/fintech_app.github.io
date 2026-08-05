#!/usr/bin/env python3
"""
Agente de controle de orçamento - sincronização com Mercado Pago.

O que faz:
  1. Busca pagamentos do Mercado Pago (API /v1/payments/search) num período.
  2. Categoriza cada pagamento usando as regras da aba "Mapeamento".
  3. Grava tudo na aba "MP_Transacoes" da planilha de orçamento.
  4. Recalcula a planilha (as abas Orcamento/Resumo_MP têm fórmulas que somam
     esses lançamentos e comparam com o Previsto automaticamente).
  5. Imprime um alerta no terminal para toda categoria que estourou o orçamento
     no mês consultado (e dispara um webhook, se configurado).

Uso:
  python3 mp_sync.py                       (sincroniza o mês atual)
  python3 mp_sync.py --mes 2025-01
  python3 mp_sync.py --mes 2025-01 --config config.json
  python3 mp_sync.py --mes 2025-01 --somente-aprovados   (padrão: True)
  python3 mp_sync.py --planilha meu_orcamento.xlsx        (sobrescreve o config.json)

A planilha não precisa ter nome nem local fixos: informe o arquivo que você
mesmo exportou/subiu via --planilha, ou defina "planilha" em config.json.
Qualquer planilha com a mesma estrutura de abas (Orcamento, Mapeamento,
MP_Transacoes, Resumo_MP) funciona — veja config.example.json.

Leitura por layout (sem Mercado Pago, sem token -- mesmo conceito do modal
"Configurar layout de leitura" do painel web, ver js/budget-ai.js):
  python3 mp_sync.py --criar-layout                        (assistente interativo, salva layout.json)
  python3 mp_sync.py --criar-layout --layout outro.json     (salva em outro caminho)
  python3 mp_sync.py --ler-orcamento --layout layout.json --planilha meu_orcamento.xlsx
                                                             (lê Previsto/Realizado com esse layout e imprime o resumo)

Requer um config.json (veja config.example.json) com:
  { "mercado_pago_access_token": "...", "planilha": "meu_orcamento.xlsx" }

Este script é pensado para rodar tanto manualmente quanto via agendamento
(sem intervenção humana): nunca lança exceção não tratada, sempre grava um
log em logs/mp_sync.log e sempre termina com uma linha final clara indicando
o resultado (OK, ESTOURADO, ou ERRO).
"""
import argparse
import json
import os
import shutil
import subprocess
import sys
import tempfile
import traceback
import unicodedata
from datetime import datetime, timedelta

try:
    import requests
except ImportError:
    sys.exit("Falta o pacote 'requests'. Rode: pip install requests --break-system-packages")

import openpyxl

import budget_layout

MP_API_URL = "https://api.mercadopago.com/v1/payments/search"

# Abas que o script lê/grava diretamente. Qualquer planilha com essa
# estrutura funciona -- não precisa ser um arquivo com nome fixo.
REQUIRED_SHEETS = ["Mapeamento", "MP_Transacoes", "Resumo_MP"]

MESES_PT = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
            "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]


def mes_label(mes_str):
    """'2025-01' -> 'Janeiro - 2025' (mesmo rótulo usado nas abas Orcamento/Resumo_MP)."""
    ano, mes = mes_str.split("-")
    return f"{MESES_PT[int(mes) - 1]} - {ano}"


def log_path_for(planilha):
    base = os.path.dirname(os.path.abspath(planilha))
    logdir = os.path.join(base, "logs")
    os.makedirs(logdir, exist_ok=True)
    return os.path.join(logdir, "mp_sync.log")


def log(planilha, linha):
    try:
        with open(log_path_for(planilha), "a", encoding="utf-8") as f:
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            f.write(f"[{ts}] {linha}\n")
    except Exception:
        pass  # log é best-effort, nunca deve derrubar o sync


def normalize(text):
    """minúsculo, sem acento, para comparação de palavra-chave."""
    if text is None:
        return ""
    text = str(text).lower().strip()
    text = unicodedata.normalize("NFKD", text)
    return "".join(c for c in text if not unicodedata.combining(c))


def month_bounds(mes_str):
    """'2025-01' -> (inicio_iso, fim_iso) cobrindo o mês inteiro, timezone -03:00."""
    ano, mes = (int(x) for x in mes_str.split("-"))
    inicio = datetime(ano, mes, 1)
    if mes == 12:
        fim = datetime(ano + 1, 1, 1)
    else:
        fim = datetime(ano, mes + 1, 1)
    fim = fim - timedelta(seconds=1)
    fmt = "%Y-%m-%dT%H:%M:%S.000-03:00"
    return inicio.strftime(fmt), fim.strftime(fmt)


def fetch_mp_payments(token, begin_date, end_date):
    """Pagina sobre /v1/payments/search e devolve a lista completa de resultados."""
    headers = {"Authorization": f"Bearer {token}"}
    results = []
    offset = 0
    limit = 50
    while True:
        params = {
            "range": "date_created",
            "begin_date": begin_date,
            "end_date": end_date,
            "sort": "date_created",
            "criteria": "desc",
            "limit": limit,
            "offset": offset,
        }
        resp = requests.get(MP_API_URL, headers=headers, params=params, timeout=30)
        if resp.status_code != 200:
            raise RuntimeError(f"Erro Mercado Pago ({resp.status_code}): {resp.text[:500]}")
        data = resp.json()
        page = data.get("results", [])
        results.extend(page)
        total = data.get("paging", {}).get("total", len(results))
        offset += limit
        if offset >= total or not page:
            break
    return results


def validate_workbook(wb, planilha):
    """Confere se a planilha enviada tem a estrutura mínima esperada, com
    mensagem clara em vez de deixar estourar um KeyError cru mais adiante."""
    faltando = [s for s in REQUIRED_SHEETS if s not in wb.sheetnames]
    if faltando:
        raise ValueError(
            f"A planilha '{planilha}' não tem a(s) aba(s) {', '.join(faltando)}. "
            "Use uma planilha de orçamento com a mesma estrutura (abas Orcamento, Mapeamento, "
            "MP_Transacoes e Resumo_MP) -- pode ser qualquer arquivo que você mesmo exportou ou "
            "subiu, não precisa ter nome nem local fixos."
        )


def load_mapping_rules(wb):
    ws = wb["Mapeamento"]
    rules = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        kw, categoria, _desc = (row + (None, None, None))[:3]
        if kw and categoria:
            rules.append((normalize(kw), categoria))
    return rules


def categorize(description, rules):
    desc_norm = normalize(description)
    for kw, categoria in rules:
        if kw and kw in desc_norm:
            return categoria
    return "Não categorizado"


def write_transactions(wb, payments, rules, mes_ref, somente_aprovados):
    ws = wb["MP_Transacoes"]

    # limpa placeholder / lançamentos já gravados para este mês (evita duplicar ao re-rodar)
    rows_to_delete = []
    for r in range(2, ws.max_row + 1):
        val_a = ws.cell(row=r, column=1).value
        val_g = ws.cell(row=r, column=7).value
        if val_a == "(vazio até rodar mp_sync.py)" or val_g == mes_ref:
            rows_to_delete.append(r)
    for r in reversed(rows_to_delete):
        ws.delete_rows(r)

    next_row = ws.max_row + 1
    gravados = 0
    for p in payments:
        status = p.get("status", "")
        if somente_aprovados and status != "approved":
            continue
        desc = p.get("description") or p.get("statement_descriptor") or "(sem descrição)"
        categoria = categorize(desc, rules)
        valor = p.get("transaction_amount", 0)
        data = (p.get("date_approved") or p.get("date_created") or "")[:10]
        ws.cell(row=next_row, column=1, value=p.get("id"))
        ws.cell(row=next_row, column=2, value=data)
        ws.cell(row=next_row, column=3, value=desc)
        ws.cell(row=next_row, column=4, value=categoria)
        ws.cell(row=next_row, column=5, value=valor)
        ws.cell(row=next_row, column=6, value=status)
        ws.cell(row=next_row, column=7, value=mes_ref)
        next_row += 1
        gravados += 1
    return gravados


def recalc(path):
    """Recalcula fórmulas via LibreOffice headless (convert-to no próprio formato).

    Best-effort: se o LibreOffice ('soffice') não estiver instalado ou não
    estiver no PATH, isso não deve derrubar o sync -- as fórmulas serão
    recalculadas normalmente quando a planilha for aberta no Excel/LibreOffice.
    """
    outdir = os.path.join(tempfile.gettempdir(), "mp_sync_recalc")
    os.makedirs(outdir, exist_ok=True)

    # 1) tenta achar no PATH; 2) senão, checa os locais padrão de instalação
    #    (Windows costuma não colocar o LibreOffice no PATH automaticamente).
    soffice = shutil.which("soffice") or shutil.which("soffice.exe")
    if not soffice:
        candidatos = [
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
            "/usr/bin/soffice",
            "/opt/libreoffice/program/soffice",
            "/Applications/LibreOffice.app/Contents/MacOS/soffice",
        ]
        soffice = next((c for c in candidatos if os.path.isfile(c)), None)

    if not soffice:
        print("[aviso] LibreOffice ('soffice') não encontrado no PATH nem nos "
              "locais padrão de instalação; pulando recálculo automático. "
              "As fórmulas serão recalculadas ao abrir a planilha no Excel/LibreOffice.")
        return
    cmd = [soffice, "--headless", "--convert-to", "xlsx", "--outdir", outdir, path]
    try:
        subprocess.run(cmd, check=True, capture_output=True, timeout=90)
        converted = os.path.join(outdir, os.path.basename(path))
        shutil.move(converted, path)
    except (subprocess.CalledProcessError, subprocess.TimeoutExpired, OSError) as e:
        print(f"[aviso] falha ao recalcular via LibreOffice ({e}); "
              "planilha gravada sem recálculo automático.")


def check_alerts(path, mes_nome):
    """Retorna (mes_existe_no_orcamento, lista_de_estouros)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Resumo_MP"]
    alerts = []
    mes_existe = False
    for row in ws.iter_rows(min_row=3, values_only=True):
        categoria, mes, previsto, gasto, saldo, status = (row + (None,) * 6)[:6]
        if mes == mes_nome:
            mes_existe = True
            if status == "ESTOURADO":
                alerts.append((categoria, previsto, gasto, saldo))
    return mes_existe, alerts


def send_webhook(url, text):
    try:
        requests.post(url, json={"text": text}, timeout=10)
    except Exception as e:
        print(f"[aviso] falha ao enviar webhook: {e}")


def _ask(prompt, default=None, required=False):
    suffix = f" [{default}]" if default not in (None, "") else ""
    while True:
        resp = input(f"{prompt}{suffix}: ").strip()
        if not resp:
            resp = default
        if resp or not required:
            return resp
        print("Esse campo é obrigatório.")


def wizard_criar_layout(dest_path):
    """Assistente interativo (equivalente em CLI ao modal 'Configurar layout
    de leitura' do painel web) que monta um layout.json passo a passo."""
    print("=== Criar layout de leitura de orçamento ===")
    print("Responda as perguntas abaixo (Enter mantém o valor padrão entre colchetes).\n")

    nome = _ask("Nome do layout", default="Meu orçamento", required=True)
    aba = _ask("Nome da aba (em branco = detectar automaticamente)", default="")

    formato = ""
    while formato not in ("longo", "largo"):
        formato = (_ask(
            "Formato: 'largo' (Previsto/Realizado por mês, lado a lado) ou "
            "'longo' (uma linha por categoria)", default="largo",
        ) or "").lower()

    layout = {"name": nome, "sheetName": aba or None, "format": formato}

    if formato == "longo":
        layout["headerRow"] = int(_ask("Linha do cabeçalho", default="1"))
        layout["colCategoria"] = _ask("Coluna da Categoria (ex: A)", default="A", required=True).upper()
        col_mes = _ask("Coluna do Mês (em branco = não usar)", default="")
        layout["colMes"] = col_mes.upper() or None
        layout["colPrevisto"] = _ask("Coluna do Previsto (ex: C)", default="C", required=True).upper()
        col_real = _ask("Coluna do Realizado (em branco = não usar)", default="")
        layout["colRealizado"] = col_real.upper() or None
    else:
        layout["colCategoriaLarga"] = _ask("Coluna da Categoria (ex: A)", default="A", required=True).upper()
        layout["monthRow"] = int(_ask("Linha com o nome dos meses", default="1"))
        layout["subHeaderRow"] = int(
            _ask("Linha com 'Previsto'/'Realizado'", default=str(layout["monthRow"] + 1))
        )

    with open(dest_path, "w", encoding="utf-8") as f:
        json.dump(layout, f, ensure_ascii=False, indent=2)
    print(f"\nLayout salvo em {dest_path}")
    print("Use com: python3 mp_sync.py --ler-orcamento --layout " + dest_path + " --planilha SUA_PLANILHA.xlsx")
    return layout


def ler_orcamento_run(args):
    """Só lê o orçamento (Previsto/Realizado por categoria/mês) usando um
    layout -- sem Mercado Pago, sem token, sem gravar nada na planilha.
    Devolve (resultado, mensagem) no mesmo estilo de run()."""
    if not args.layout:
        return "erro", "Informe --layout caminho/para/layout.json (crie um com --criar-layout)."
    if not os.path.exists(args.layout):
        return "erro", f"Layout não encontrado: {args.layout}"
    with open(args.layout, encoding="utf-8") as f:
        layout = json.load(f)

    planilha = args.planilha
    if not planilha and os.path.exists(args.config):
        with open(args.config, encoding="utf-8") as f:
            cfg = json.load(f)
        planilha = cfg.get("planilha")
    if not planilha:
        return "erro", (
            "Informe --planilha caminho/para/seu_orcamento.xlsx (ou defina \"planilha\" em config.json)."
        )
    if not os.path.exists(planilha):
        return "erro", f"Planilha não encontrada: {planilha}"

    wb = openpyxl.load_workbook(planilha, data_only=True)
    try:
        resultado = budget_layout.analyze_with_layout(wb, layout)
    except ValueError as e:
        return "erro", str(e)

    nome_layout = layout.get("name") or args.layout
    linhas = [f"Leitura de '{planilha}' (aba \"{resultado['sheetName']}\", layout \"{nome_layout}\"):"]
    for r in resultado["rows"]:
        rotulo = r["categoria"] + (f" ({r['mes']})" if r["mes"] else "")
        linhas.append(
            f"  - {rotulo}: previsto R$ {r['previsto']:,.2f} | realizado R$ {r['realizado']:,.2f} | "
            f"saldo R$ {r['saldo']:,.2f} | {r['status']}"
        )
    linhas.append(
        f"Total previsto: R$ {resultado['totalPrevisto']:,.2f} | "
        f"Total realizado: R$ {resultado['totalRealizado']:,.2f} | "
        f"Saldo: R$ {resultado['saldoTotal']:,.2f}"
    )

    if resultado["overBudget"]:
        linhas.append(f"\n⚠️  {len(resultado['alerts'])} categoria(s) estouraram o orçamento.")
        return "estourado", "\n".join(linhas)

    linhas.append("\n✅ Nenhuma categoria estourou o orçamento.")
    return "ok", "\n".join(linhas)


def run(args):
    """Executa o sync e devolve (resultado, mensagem) para uso programático e pelo agendamento."""
    if not os.path.exists(args.config):
        return "erro", (f"Config não encontrado: {args.config}. "
                         f"Copie config.example.json -> config.json e preencha o token.")

    with open(args.config, encoding="utf-8") as f:
        cfg = json.load(f)

    token = cfg.get("mercado_pago_access_token")
    if not token or token.startswith("COLE_"):
        return "erro", "Access token do Mercado Pago não configurado em config.json."

    # --planilha (CLI) tem prioridade sobre "planilha" no config.json -- assim
    # dá pra apontar para qualquer arquivo que você mesmo subiu, sem precisar
    # editar o config.json nem usar um nome fixo.
    planilha = getattr(args, "planilha", None) or cfg.get("planilha")
    if not planilha:
        return "erro", (
            "Nenhuma planilha de orçamento informada. Use --planilha caminho/para/seu_orcamento.xlsx "
            "ou defina \"planilha\" em config.json (veja config.example.json)."
        )
    config_dir = os.path.dirname(os.path.abspath(args.config))
    if not os.path.isabs(planilha):
        planilha = os.path.join(config_dir, planilha)
    if not os.path.exists(planilha):
        return "erro", f"Planilha não encontrada: {planilha}"

    mes = args.mes or datetime.now().strftime("%Y-%m")
    mes_nome = mes_label(mes)

    begin, end = month_bounds(mes)
    print(f"Buscando pagamentos do Mercado Pago entre {begin} e {end} ...")
    payments = fetch_mp_payments(token, begin, end)
    print(f"{len(payments)} pagamento(s) encontrado(s) no período.")

    wb = openpyxl.load_workbook(planilha)
    try:
        validate_workbook(wb, planilha)
    except ValueError as e:
        return "erro", str(e)
    rules = load_mapping_rules(wb)
    gravados = write_transactions(wb, payments, rules, mes, args.somente_aprovados)
    wb.save(planilha)
    print(f"{gravados} lançamento(s) gravado(s) em MP_Transacoes (status filtrado: aprovados={args.somente_aprovados}).")

    print("Recalculando planilha...")
    recalc(planilha)

    mes_existe, alerts = check_alerts(planilha, mes_nome)
    log(planilha, f"mes={mes} gravados={gravados} mes_existe_no_orcamento={mes_existe} estouros={len(alerts)}")

    if not mes_existe:
        msg = (f"O mês {mes_nome} ainda não tem colunas Previsto/Realizado na aba Orcamento. "
               f"{gravados} pagamento(s) foram gravados em MP_Transacoes mesmo assim, mas não dá "
               f"para comparar com orçamento até você adicionar esse mês na planilha.")
        print(f"\n⚠️  {msg}")
        return "sem_orcamento", msg

    if alerts:
        linhas = []
        for categoria, previsto, gasto, saldo in alerts:
            linha = f"  - {categoria}: previsto R$ {previsto:,.2f} | gasto MP R$ {gasto:,.2f} | excesso R$ {abs(saldo):,.2f}"
            print(linha)
            linhas.append(linha)
        msg = f"Orçamento ESTOURADO em {mes_nome}:\n" + "\n".join(linhas)
        print(f"\n⚠️  {msg}")
        webhook = cfg.get("alerta_webhook")
        if webhook:
            send_webhook(webhook, msg)
        return "estourado", msg

    msg = f"Nenhuma categoria estourou o orçamento em {mes_nome} ({gravados} pagamento(s) sincronizado(s))."
    print(f"\n✅ {msg}")
    return "ok", msg


def main():
    ap = argparse.ArgumentParser(description="Sincroniza pagamentos do Mercado Pago com a planilha de orçamento.")
    ap.add_argument("--mes", default=None, help="Mês a sincronizar, formato AAAA-MM (padrão: mês atual)")
    ap.add_argument("--config", default="config.json", help="Caminho do config.json")
    ap.add_argument(
        "--planilha", default=None,
        help="Caminho da planilha de orçamento a sincronizar (sobrescreve \"planilha\" do config.json). "
             "Aceita qualquer arquivo com a mesma estrutura de abas, sem nome nem local fixos.",
    )
    ap.add_argument("--somente-aprovados", dest="somente_aprovados", action="store_true", default=True)
    ap.add_argument("--incluir-todos-status", dest="somente_aprovados", action="store_false")
    ap.add_argument(
        "--layout", default=None,
        help="Caminho de um layout.json (ver layout.example.json) que descreve como ler a aba de "
             "orçamento -- mesmo conceito do modal 'Configurar layout de leitura' do painel web.",
    )
    ap.add_argument(
        "--criar-layout", dest="criar_layout", action="store_true",
        help="Assistente interativo para criar um layout.json (não sincroniza nada, só cria o arquivo "
             "em --layout, ou layout.json por padrão).",
    )
    ap.add_argument(
        "--ler-orcamento", dest="ler_orcamento", action="store_true",
        help="Só lê o orçamento (Previsto/Realizado por categoria/mês) usando --layout e imprime o "
             "resumo -- não mexe no Mercado Pago, não precisa de token, não grava nada na planilha.",
    )
    args = ap.parse_args()

    if args.criar_layout:
        try:
            wizard_criar_layout(args.layout or "layout.json")
        except (KeyboardInterrupt, EOFError):
            print("\nCancelado.")
            sys.exit(1)
        return

    if args.ler_orcamento:
        try:
            resultado, msg = ler_orcamento_run(args)
        except Exception:
            print("\n❌ ERRO ao ler o orçamento:")
            traceback.print_exc()
            sys.exit(1)
        print(msg)
        sys.exit(0 if resultado == "ok" else (2 if resultado == "estourado" else 1))

    try:
        resultado, msg = run(args)
    except Exception:
        print("\n❌ ERRO ao sincronizar com o Mercado Pago:")
        traceback.print_exc()
        try:
            planilha_hint = args.config
            log(planilha_hint, f"ERRO: {traceback.format_exc()}")
        except Exception:
            pass
        sys.exit(1)

    sys.exit(0 if resultado in ("ok",) else (2 if resultado == "estourado" else (3 if resultado == "sem_orcamento" else 1)))


if __name__ == "__main__":
    main()
