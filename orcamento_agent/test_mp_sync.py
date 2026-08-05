"""
Teste ponta-a-ponta do mp_sync.py sem chamar a API real do Mercado Pago.
Cobre: categorização, gravação em MP_Transacoes, recálculo, detecção de
estouro, e o caso de sincronizar um mês que ainda não existe no orçamento.

A planilha usada aqui é montada na hora por build_fixture_workbook(), com
dados fictícios -- o teste não depende de nenhum orçamento real do usuário
nem de um arquivo fixo no disco.
"""
import json
import argparse
import openpyxl
import mp_sync

TEST_PLANILHA = "TEST_Orcamento.xlsx"


def build_fixture_workbook(path):
    """Monta uma planilha mínima com a mesma estrutura exigida por
    mp_sync.py (abas Orcamento, Mapeamento, MP_Transacoes, Resumo_MP) —
    dados fictícios só para exercitar o script."""
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Orcamento"
    ws.append([None, "ORÇAMENTO DE TESTE"])
    ws.append([None, "Descrição", "Forma Pagamento", "Categoria", "Status", "Total Ano", "Total Pago", "Janeiro - 2025", None])
    ws.append([None, None, None, None, None, None, None, "Previsto", "Realizado"])
    ws.append([None, "Academia", "Cartão", "Saúde", "Recorrente", 18000, 0, 1500, 0])
    ws.append([None, "Limpeza", "Pix", "Moradia", "Recorrente", 19200, 0, 1600, 0])
    ws.merge_cells("H2:I2")

    ws2 = wb.create_sheet("Mapeamento")
    ws2.append(["Legenda: edite a coluna 'Palavra-chave' para casar com a descrição do pagamento no Mercado Pago."])
    ws2.append(["Palavra-chave (contém, sem acento, minúsculo)", "Categoria", "Descrição dos Gastos (opcional)"])
    ws2.append(["gympass", "Saúde", "Academia"])
    ws2.append(["faxineira", "Moradia", "Limpeza"])

    ws3 = wb.create_sheet("MP_Transacoes")
    ws3.append(["ID Pagamento MP", "Data", "Descrição", "Categoria (mapeada)", "Valor (R$)", "Status MP", "Mês Referência"])

    ws4 = wb.create_sheet("Resumo_MP")
    ws4.append(["RESUMO: Orçamento (Previsto) x Gasto real no Mercado Pago"])
    ws4.append(["Categoria", "Mês", "Previsto (soma)", "Gasto Mercado Pago", "Saldo", "Status"])
    ws4.append([
        "Saúde", "Janeiro - 2025",
        "=SUMIFS(Orcamento!H4:H5,Orcamento!D4:D5,A3)",
        '=SUMIFS(MP_Transacoes!E:E,MP_Transacoes!D:D,A3,MP_Transacoes!G:G,"2025-01")',
        "=C3-D3",
        '=IF(D3=0,"sem dados MP",IF(D3>C3,"ESTOURADO","DENTRO DO ORÇAMENTO"))',
    ])
    ws4.append([
        "Moradia", "Janeiro - 2025",
        "=SUMIFS(Orcamento!H4:H5,Orcamento!D4:D5,A4)",
        '=SUMIFS(MP_Transacoes!E:E,MP_Transacoes!D:D,A4,MP_Transacoes!G:G,"2025-01")',
        "=C4-D4",
        '=IF(D4=0,"sem dados MP",IF(D4>C4,"ESTOURADO","DENTRO DO ORÇAMENTO"))',
    ])

    wb.save(path)


build_fixture_workbook(TEST_PLANILHA)

# Pagamentos simulados de Janeiro/2025: Gympass estourando o previsto de R$1.500
fake_payments = [
    {"id": 111, "description": "GYMPASS ASSINATURA MENSAL", "transaction_amount": 1500.00,
     "status": "approved", "date_approved": "2025-01-05T10:00:00.000-03:00"},
    {"id": 112, "description": "GYMPASS TAXA EXTRA", "transaction_amount": 400.00,
     "status": "approved", "date_approved": "2025-01-20T10:00:00.000-03:00"},
    {"id": 113, "description": "PIX FAXINEIRA MARIA", "transaction_amount": 1600.00,
     "status": "approved", "date_approved": "2025-01-10T10:00:00.000-03:00"},
    {"id": 114, "description": "PAGAMENTO NAO MAPEADO XYZ", "transaction_amount": 50.00,
     "status": "approved", "date_approved": "2025-01-15T10:00:00.000-03:00"},
    {"id": 115, "description": "GYMPASS PENDENTE", "transaction_amount": 999.00,
     "status": "rejected", "date_approved": "2025-01-16T10:00:00.000-03:00"},
]

wb = openpyxl.load_workbook(TEST_PLANILHA)
rules = mp_sync.load_mapping_rules(wb)

# categorização
cat1 = mp_sync.categorize("GYMPASS ASSINATURA MENSAL", rules)
cat2 = mp_sync.categorize("PIX FAXINEIRA MARIA", rules)
cat3 = mp_sync.categorize("PAGAMENTO NAO MAPEADO XYZ", rules)
assert cat1 == "Saúde", f"esperado Saúde, veio {cat1}"
assert cat2 == "Moradia", f"esperado Moradia, veio {cat2}"
assert cat3 == "Não categorizado", f"esperado Não categorizado, veio {cat3}"
print("OK categorização:", cat1, "|", cat2, "|", cat3)

# mes_label
assert mp_sync.mes_label("2025-01") == "Janeiro - 2025"
assert mp_sync.mes_label("2026-08") == "Agosto - 2026"
print("OK mes_label")

# validação de estrutura (abas obrigatórias presentes)
mp_sync.validate_workbook(wb, TEST_PLANILHA)
print("OK validate_workbook aceita a planilha de teste")

gravados = mp_sync.write_transactions(wb, fake_payments, rules, "2025-01", somente_aprovados=True)
wb.save(TEST_PLANILHA)
assert gravados == 4, gravados
print(f"OK gravados: {gravados}")

mp_sync.recalc(TEST_PLANILHA)

mes_existe, alerts = mp_sync.check_alerts(TEST_PLANILHA, "Janeiro - 2025")
assert mes_existe is True
categorias_estouradas = {a[0] for a in alerts}
assert "Saúde" in categorias_estouradas, alerts
assert "Moradia" not in categorias_estouradas, alerts
print("OK alertas Janeiro-2025:", alerts)

# mês que NÃO existe na planilha (ex: mês atual real, fora do orçamento de teste)
mes_existe_fora, alerts_fora = mp_sync.check_alerts(TEST_PLANILHA, "Agosto - 2026")
assert mes_existe_fora is False, "Agosto-2026 não deveria existir nesse orçamento de teste"
assert alerts_fora == []
print("OK detecção de mês fora do orçamento (não finge que está tudo certo)")

# valores exatos na aba Resumo_MP
wb2 = openpyxl.load_workbook(TEST_PLANILHA, data_only=True)
ws2 = wb2["Resumo_MP"]
found = {}
for row in ws2.iter_rows(min_row=3, values_only=True):
    categoria, mes, previsto, gasto, saldo, status = (row + (None,) * 6)[:6]
    if mes == "Janeiro - 2025":
        found[categoria] = (previsto, gasto, saldo, status)

assert found["Saúde"][1] == 1900.00, found["Saúde"]
assert found["Saúde"][3] == "ESTOURADO"
assert found["Moradia"][1] == 1600.00
assert found["Moradia"][3] == "DENTRO DO ORÇAMENTO"
print("OK valores Resumo_MP:", found["Saúde"], "|", found["Moradia"])

# MP_Transacoes sem o pagamento rejeitado
ws3 = openpyxl.load_workbook(TEST_PLANILHA)["MP_Transacoes"]
descs = [r[2] for r in ws3.iter_rows(min_row=2, values_only=True) if r[0]]
assert len(descs) == 4
assert "GYMPASS PENDENTE" not in descs
print("OK MP_Transacoes:", descs)

# planilha sem as abas esperadas -> validate_workbook recusa com mensagem clara
wb_incompleta = openpyxl.Workbook()
wb_incompleta.active.title = "SomenteUmaAba"
try:
    mp_sync.validate_workbook(wb_incompleta, "planilha_incompleta.xlsx")
    raise AssertionError("validate_workbook deveria ter recusado a planilha incompleta")
except ValueError as e:
    assert "Mapeamento" in str(e) and "MP_Transacoes" in str(e) and "Resumo_MP" in str(e)
print("OK validate_workbook recusa planilha sem a estrutura esperada")

# roda run() completo com um config.json falso apontando para a planilha de teste
with open("TEST_config.json", "w", encoding="utf-8") as f:
    json.dump({"mercado_pago_access_token": "TEST-fake", "planilha": TEST_PLANILHA}, f)

Args = argparse.Namespace(mes="2025-01", config="TEST_config.json", somente_aprovados=True, planilha=None)

# monkeypatch fetch_mp_payments pra não bater na API real
mp_sync.fetch_mp_payments = lambda token, begin, end: fake_payments
resultado, msg = mp_sync.run(Args)
assert resultado == "estourado", (resultado, msg)
print("OK run() end-to-end:", resultado)

# --planilha (CLI) sobrescreve o que está no config.json
with open("TEST_config_sem_planilha.json", "w", encoding="utf-8") as f:
    json.dump({"mercado_pago_access_token": "TEST-fake"}, f)
Args2 = argparse.Namespace(mes="2025-01", config="TEST_config_sem_planilha.json", somente_aprovados=True, planilha=TEST_PLANILHA)
resultado2, msg2 = mp_sync.run(Args2)
assert resultado2 == "estourado", (resultado2, msg2)
print("OK --planilha via CLI sobrescreve config.json:", resultado2)

# nenhuma planilha configurada (nem CLI, nem config.json) -> erro claro
Args3 = argparse.Namespace(mes="2025-01", config="TEST_config_sem_planilha.json", somente_aprovados=True, planilha=None)
resultado3, msg3 = mp_sync.run(Args3)
assert resultado3 == "erro" and "planilha" in msg3.lower(), (resultado3, msg3)
print("OK erro claro quando nenhuma planilha é informada")

print("\nTODOS OS TESTES PASSARAM ✅")
